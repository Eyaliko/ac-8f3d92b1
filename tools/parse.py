#!/usr/bin/env python3
import os, re, json, csv, datetime as dt, warnings, collections
import openpyxl
warnings.filterwarnings("ignore")

DATE_RE = re.compile(r"(\d{1,2})/(\d{1,2})/(\d{4})")

def s(v):
    return "" if v is None else str(v).strip()

def num(v):
    if v is None or v == "": return None
    if isinstance(v, (int, float)): return float(v)
    t = str(v).replace(",", "").replace("€", "").strip().rstrip("%")
    try: return float(t)
    except ValueError: return None

def dmy(text):
    m = DATE_RE.search(text or "")
    if not m: return None
    d, mo, y = map(int, m.groups())
    try: return dt.date(y, mo, d)
    except ValueError: return None

def anydate(v):
    if isinstance(v, dt.datetime): return v.date()
    if isinstance(v, dt.date): return v
    t = s(v)
    if not t: return None
    d = dmy(t)
    if d: return d
    for f in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y"):
        try: return dt.datetime.strptime(t, f).date()
        except ValueError: pass
    return None

def kind(fn):
    f = fn.lower()
    if "manager" in f: return "manager"
    if "booking_plan" in f or "bookings_plan" in f: return "bookplan"
    if "accommodation" in f or "reservations_list" in f: return "accom"
    return None

# ---------------- manager report ----------------
MGR_SECTIONS = {
    "Rooms": "rooms", "Guests": "guests", "Income": "income",
    "Average consumption Per Occupied Room": "arr",
    "Average Consumption per Available Room": "apar",
    "Average Consumption Per Guest": "apg",
}
def slug(x):
    return re.sub(r"[^a-z0-9]+", "_", x.lower()).strip("_")

def parse_manager(path, folder_day):
    ws = openpyxl.load_workbook(path, data_only=True).active
    rows = list(ws.iter_rows(values_only=True))
    created = report_date = None
    for r in rows[:4]:
        joined = " | ".join(s(c) for c in r)
        if "Created:" in joined:
            created = dmy(s(r[1]))
            for i, c in enumerate(r):
                if s(c).lower().startswith("report date") and i + 1 < len(r):
                    report_date = anydate(r[i + 1])
    if report_date is None: report_date = created - dt.timedelta(days=1) if created else None
    if report_date is None: return None
    # detect prior-year block
    yearrow = None
    for r in rows[:5]:
        cells = [s(c) for c in r]
        if any(re.fullmatch(r"20\d\d", c) for c in cells): yearrow = cells; break
    cy = py = None
    if yearrow:
        ys = [(i, int(c)) for i, c in enumerate(yearrow) if re.fullmatch(r"20\d\d", c)]
        if ys: cy = ys[0][1]
        if len(ys) > 1: py = ys[1][1]
    rec = {"report_date": report_date.isoformat(), "created": created.isoformat() if created else "",
           "source_day": folder_day, "cur_year": cy, "prior_year": py}
    sec = ""
    for r in rows:
        a = s(r[0])
        if not a: continue
        vals = [v for v in r[1:8]]
        if all(v is None or s(v) == "" for v in vals):
            if a in MGR_SECTIONS: sec = MGR_SECTIONS[a]
            continue
        if not sec: continue
        base = f"{sec}_{slug(a)}"
        for j, per in enumerate(["day", "mtd", "ytd"]):
            rec[f"{base}_{per}"] = num(vals[j]) if j < len(vals) else None
        if py:
            for j, per in enumerate(["day", "mtd", "ytd"]):
                k = j + 3
                rec[f"{base}_ly_{per}"] = num(vals[k]) if k < len(vals) else None
    return rec

# ---------------- booking plan ----------------
BP_ROWS = {"%": "pct", "Booking position": "booking_position", "Allotments (Deduct)": "allotments",
           "Occupancy %": "occupancy_pct", "Physically available": "physically_available",
           "Actual rooms": "actual_rooms", "Rooms to sell": "rooms_to_sell", "Day use": "day_use",
           "Total": "total"}
NON_ROOMTYPE = set(BP_ROWS) | {"Room types", "Metric", "Totals"}

def parse_bookplan(path, folder_day):
    ws = openpyxl.load_workbook(path, data_only=True).active
    rows = list(ws.iter_rows(values_only=True))
    created = from_d = to_d = None
    for r in rows[:4]:
        j = " | ".join(s(c) for c in r)
        if "Created:" in j:
            created = dmy(s(r[1]))
            for i, c in enumerate(r):
                lc = s(c).lower()
                if lc == "from date" and i + 1 < len(r): from_d = anydate(r[i + 1])
                if lc == "to date" and i + 1 < len(r): to_d = anydate(r[i + 1])
    hi = next((i for i, r in enumerate(rows) if s(r[0]) == "Metric"), None)
    if hi is None or from_d is None: return []
    if created is None: created = anydate(folder_day)
    hdr = rows[hi]
    cols = []           # (col_index, stay_date)
    y, m0 = from_d.year, from_d.month
    prev = None
    for i in range(2, len(hdr)):
        c = hdr[i]
        if isinstance(c, (dt.datetime, dt.date)):
            cols.append((i, anydate(c))); prev = cols[-1][1]; continue
        t = s(c)
        mm = re.fullmatch(r"(\d{1,2})/(\d{1,2})", t)
        if not mm: continue
        d, mo = int(mm.group(1)), int(mm.group(2))
        yy = y if mo >= m0 else y + 1
        try: sd = dt.date(yy, mo, d)
        except ValueError: continue
        if prev and sd < prev: sd = dt.date(yy + 1, mo, d)
        cols.append((i, sd)); prev = sd
    if not cols: return []
    grid = collections.defaultdict(dict)   # stay_date -> {metric: val}
    caps = {}
    seen_total = 0
    for r in rows[hi + 1:]:
        a = s(r[0])
        if not a or a == "Room types": continue
        if a == "Total":
            seen_total += 1
            if seen_total > 1: continue
            key = "rooms_sold"
        elif a in BP_ROWS:
            key = BP_ROWS[a]
        else:
            key = "rt::" + a
            caps[a] = num(r[1])
        for ci, sd in cols:
            v = num(r[ci]) if ci < len(r) else None
            if v is not None: grid[sd][key] = v
    out = []
    for sd, met in sorted(grid.items()):
        rec = {"snapshot_date": created.isoformat(), "stay_date": sd.isoformat(),
               "lead_days": (sd - created).days, "source_day": folder_day}
        rt = {}
        for k, v in met.items():
            if k.startswith("rt::"): rt[k[4:]] = v
            else: rec[k] = v
        rec["by_room_type"] = rt
        out.append(rec)
    return out

# ---------------- accommodations ----------------
# NOTE: "Full Name" and "Email" are deliberately NOT mapped. This database is
# published to a public repository; guest identities must never enter it.
A_MAP = {"Reservation ID": "reservation_id", "Book Date": "book_date",
         "Room": "room", "Room Type": "room_type", "Check-in": "check_in", "Check-out": "check_out",
         "Nights": "nights", "Adults": "adults", "Children": "children", "Infants": "infants",
         "Rate": "rate_plan", "Total": "total", "Room Total": "room_total", "Board": "board",
         "Balance": "balance", "Source": "source", "Commission": "commission",
         "Status": "status", "Country": "country", "Property ID": "property_id"}
NUMF = {"nights", "adults", "children", "infants", "total", "room_total", "balance", "commission"}
DATEF = {"book_date", "check_in", "check_out"}


B_MAP = {"Reservation ID": "reservation_id", "Book Date": "book_date",
         "From": "check_in", "To": "check_out", "Total": "total", "Rooms Total": "room_total",
         "Balance": "balance", "Status": "status", "Source": "source",
         "Extras total": "extras_total", "Fees total": "fees_total", "Products total": "products_total",
         "F&B total": "fb_total", "Other total": "other_total", "Segment": "segment"}
B_NUM = {"total", "room_total", "balance", "extras_total", "fees_total", "products_total", "fb_total", "other_total"}

def parse_accom_b(rows, hdr, folder_day):
    idx = {h: i for i, h in enumerate(hdr)}
    out = []
    for r in rows[1:]:
        if all(c is None or s(c) == "" for c in r): continue
        rec = {"snapshot_date": folder_day, "variant": "b", "room_type": "", "country": "", "room": ""}
        for h, key in B_MAP.items():
            if h not in idx: continue
            v = r[idx[h]] if idx[h] < len(r) else None
            if key in B_NUM: rec[key] = num(v)
            elif key in DATEF:
                d = anydate(v); rec[key] = d.isoformat() if d else ""
            else: rec[key] = s(v)
        if not rec.get("reservation_id"): continue
        try:
            ci = dt.date.fromisoformat(rec["check_in"]); co = dt.date.fromisoformat(rec["check_out"])
            rec["nights"] = float((co - ci).days)
        except Exception: rec["nights"] = None
        out.append(rec)
    return out

def parse_accom(path, folder_day):
    ws = openpyxl.load_workbook(path, data_only=True).active
    rows = list(ws.iter_rows(values_only=True))
    if not rows: return [], "empty"
    hdr = [s(c) for c in rows[0]]
    if "Room Type" not in hdr or "Check-in" not in hdr:
        return parse_accom_b(rows, hdr, folder_day), "variant_b_parsed"
    idx = {h: i for i, h in enumerate(hdr)}
    out = []
    for r in rows[1:]:
        if all(c is None or s(c) == "" for c in r): continue
        rec = {"snapshot_date": folder_day, "variant": "a"}
        for h, key in A_MAP.items():
            if h not in idx: continue
            v = r[idx[h]] if idx[h] < len(r) else None
            if key in NUMF: rec[key] = num(v)
            elif key in DATEF:
                d = anydate(v); rec[key] = d.isoformat() if d else ""
            else: rec[key] = s(v)
        if not rec.get("reservation_id"): continue
        out.append(rec)
    return out, "ok"

# ---------------- driver ----------------
import pipeline as P

def parse_dir(raw_root, only_days=None):
    """Parse day folders under raw_root. Returns (daily[], pace[], res[], stats)."""
    daily, pace, res = [], [], []
    stats = collections.Counter()
    days = sorted(os.listdir(raw_root))
    if only_days: days = [d for d in days if d in set(only_days)]
    for day in days:
        dpath = os.path.join(raw_root, day)
        if not os.path.isdir(dpath): continue
        for fn in sorted(os.listdir(dpath)):
            if not fn.lower().endswith((".xlsx", ".xls")): continue
            k = kind(fn); p = os.path.join(dpath, fn)
            try:
                if k == "manager":
                    r = parse_manager(p, day)
                    if r: daily.append(r); stats["manager"] += 1
                elif k == "bookplan":
                    rs = parse_bookplan(p, day); pace.extend(rs); stats["bookplan"] += 1
                elif k == "accom":
                    rs, status = parse_accom(p, day); res.extend(rs); stats["accom_" + status] += 1
                else:
                    stats["unknown"] += 1
            except Exception as e:
                stats["error"] += 1
                print("PARSE ERROR", day, fn, type(e).__name__, str(e)[:100], file=sys.stderr)
    return daily, pace, res, stats

def main():
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument("--raw", default=os.path.join(P.REPO, "raw"))
    ap.add_argument("--days", nargs="*", help="only these day folders (YYYY-MM-DD)")
    a = ap.parse_args()
    daily, pace, res, stats = parse_dir(a.raw, a.days)
    print("parsed:", dict(stats))
    for store, recs in (("daily", daily), ("pace", pace), ("res", res)):
        added, total = P.merge(store, recs)
        print(f"  {store:6s} +{added:6d} new, {total:7d} total")

if __name__ == "__main__":
    main()
